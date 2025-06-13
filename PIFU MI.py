# Databricks notebook source
from env import env
from src import utils, excel

import openpyxl
import pandas as pd

from pyspark.sql import functions as F
from datetime import datetime
from openpyxl.styles import NamedStyle

# COMMAND ----------

#Load PIFU data
df_raw_pifu = spark.read.option("header","true").option("recursiveFileLookup","true").parquet(env["pifu_path"])
wb = openpyxl.load_workbook('report_template.xlsx')
report_start = 'August 2021 to '

#display (df_raw_pifu)
df_raw_pifu.display()

publishing_month = df_raw_pifu.select(F.max("EROC_DerMonth")).collect()[0][0]
publishing_month = datetime.strptime(publishing_month, '%Y-%m-%d')
publishing_month = publishing_month.strftime("%B %Y")
date_header = (report_start + publishing_month) 

# COMMAND ----------

# MAGIC %md
# MAGIC #specialty

# COMMAND ----------

#creating the specialty sheet data frame
df_tfc_pifu = (df_raw_pifu

  .where(F.col("EROC_DerMetricReportingName") == "Moved and Discharged") 
  .where(F.col("EROC_DerMonth") > "2021-03-01")   
  .groupBy(
    
    "RTT_Specialty_code",
    "RTT_Specialty_Description",
    "EROC_DerMonth" )  
  .agg(F.sum("EROC_Value").alias("Moved_or_Discharged") )  
  .orderBy("EROC_DerMonth", "RTT_Specialty_code") 
  .select("EROC_DerMonth", "RTT_Specialty_code", "RTT_Specialty_Description", "Moved_or_Discharged") 
  .withColumn("RTT_Specialty_Description", F.regexp_replace("RTT_Specialty_Description","â€“", "-")) 
)

display(df_tfc_pifu)

# COMMAND ----------

#pivoting the specialty sheet data frame
df_tfc_pivot = (df_tfc_pifu
                .groupBy("RTT_Specialty_code", "RTT_Specialty_Description")
.pivot("EROC_DerMonth")
.agg(F.sum("Moved_or_Discharged") )
.orderBy("RTT_Specialty_code")
.withColumnRenamed("RTT_Specialty_code", "RTT Specialty Code")
.withColumnRenamed("RTT_Specialty_Description", "RTT Specialty Description")
)
display(df_tfc_pivot)

# COMMAND ----------

#creating enland totals 
#pivot must come straight after groupby
df_England_pivot = (df_tfc_pifu
    .groupBy()
    .pivot("EROC_DerMonth")
    .agg(F.sum("Moved_or_Discharged") )
)

#inserting into pandas
#only use toPandas on final data frame
df_England_pivot_pd = df_England_pivot.toPandas()

display(df_England_pivot)

# COMMAND ----------


#creating a loop to copy date format across to new and existing data columns
for column in df_tfc_pivot.columns[2:]:
    month_format = datetime.strptime(column, '%Y-%m-%d')
    month_format = month_format.strftime("%b-%Y") 
    df_tfc_pivot = df_tfc_pivot.withColumnRenamed(column, month_format)



# COMMAND ----------

#converting the moved and discharged data inot pandas 
df_tfc_pivot_pd = df_tfc_pivot.toPandas()

# COMMAND ----------


#inserting pandas dataframe into excel sheet 
ws_speciality = wb['PIFU | England & Specialty']
excel.insert_pandas_df_into_excel(
    df = df_tfc_pivot_pd,
    ws = ws_speciality,
    header = True,
    startrow = 11,
    startcol = 2,
    index = False,
)

#instering england totals 

excel.insert_pandas_df_into_excel(
    df= df_England_pivot_pd,
    ws = ws_speciality,
    header = False,
    startrow = 36,
    startcol = 4,
    index = False,
)


# COMMAND ----------

#adding date header (in description)
ws_speciality.cell(row=3, column=3).value = date_header

# COMMAND ----------

#defining boundaries of cells being copied 
number_of_months = df_tfc_pifu.select(F.countDistinct("EROC_DerMonth")).collect()[0][0]
new_months = number_of_months - 38
pre_date_columns = 3
copy_column = 38 + pre_date_columns
end_column = number_of_months + pre_date_columns + 1

# COMMAND ----------

print(number_of_months)

# COMMAND ----------

#copying styles from template to new data
for column_number in range(copy_column  , end_column):
    for row_number in range(11, 37):
        cell_to_copy_from = ws_speciality.cell(row=row_number, column=copy_column)
        cell_to_copy_to = ws_speciality.cell(row=row_number, column=column_number)
        excel.copy_all_cell_styles(cell_to_copy_from, cell_to_copy_to)


# COMMAND ----------

# Define the number format style
number_style = NamedStyle(name="number", number_format="0")

# Apply the number format to the specified range
for row in ws_speciality.iter_rows(min_row=12, max_row=36, min_col=pre_date_columns + 1, max_col=end_column):
    for cell in row:
        cell.number_format = number_style.number_format

# COMMAND ----------

# MAGIC %md
# MAGIC #month and org

# COMMAND ----------

df_processed_pifu = (df_raw_pifu
    .where ( F.col("EROC_DerMetricReportingName") == "Moved and Discharged")
    .where ( F.col("EROC_DerMonth") > '2021-03-01')
    .groupby(
        "EROC_DerMonth",
	    "EROC_DerProviderCode",
        "EROC_DerProviderName",
        "EROC_DerRegionName",
        "EROC_DerRegionCode",
        "EROC_DerICBCode",
        "EROC_DerICBName",
        "EROC_DerProviderAcuteStatus"
    )
    .agg( F.sum("EROC_Value").alias("Moved_or_Discharged"))
    .orderBy(
        "EROC_DerMonth",
        "EROC_DerProviderCode"
    )
)
       

#   --For ICB/Region filters, remove the filter for 'Acute' providers, and include the relevant ICB/Region fields for aggregation. 

display (df_processed_pifu)

# COMMAND ----------

#putting the data in pivot table format
df_provider_pivot = (df_processed_pifu
    .groupby(
        "EROC_DerRegionCode",
        "EROC_DerRegionName",
        "EROC_DerICBCode",
        "EROC_DerICBName",
        "EROC_DerProviderCode",
        "EROC_DerProviderName",
        "EROC_DerProviderAcuteStatus",
    )
    .pivot("EROC_DerMonth")
    .agg(F.sum("Moved_or_Discharged"))
    .orderBy(
        "EROC_DerRegionName",
        "EROC_DerICBCode",
        "EROC_DerProviderCode"
        )
    .withColumnRenamed("EROC_DerRegionCode", "Region Code")
    .withColumnRenamed("EROC_DerRegionName", "Region Name")
    .withColumnRenamed("EROC_DerICBCode", "ICB Code")
    .withColumnRenamed("EROC_DerICBName", "ICB Name")
    .withColumnRenamed("EROC_DerProviderCode", "Provider Code")
    .withColumnRenamed("EROC_DerProviderName", "Provider Name")
    .withColumnRenamed("EROC_DerProviderAcuteStatus", "Acute Status")
)

display(df_provider_pivot)

# COMMAND ----------

#formatting date headers along the pivot
for column in df_provider_pivot.columns[7:]:
    month_format = datetime.strptime(column, '%Y-%m-%d')
    month_format = month_format.strftime("%b-%Y")
    df_provider_pivot = df_provider_pivot.withColumnRenamed(column, month_format)
print(df_provider_pivot.columns)


# COMMAND ----------

#converting the pivot to pandas databframe

df_pd_provider_pivot = df_provider_pivot.toPandas()

# COMMAND ----------

#creating a workbook

ws_provider = wb['PIFU | By Org & Month']

excel.insert_pandas_df_into_excel(
    df = df_pd_provider_pivot,
    ws = wb['PIFU | By Org & Month'],
    header = True,
    startrow = 11,
    startcol = 2,
    index = False,
)

# Check width of table and get column numbers of unfomatted columns 
number_of_months = df_processed_pifu.select("EROC_DerMonth").distinct().count()
new_months = number_of_months - 42
pre_date_columns = 8 
copy_column = 42 + pre_date_columns
end_column = number_of_months + pre_date_columns + 1

# copy and paste formatting onto unformatted columns
for column_number in range (copy_column, end_column):
    for row_number in range(11,154):
        cell_to_copy_from = ws_provider.cell(row=row_number, column=copy_column)
        cell_to_paste_to = ws_provider.cell(row=row_number, column=column_number)
        excel.copy_all_cell_styles(cell_to_copy_from, cell_to_paste_to)

# get height of table and get row numbers of unformatted rows
number_of_providers = df_processed_pifu.select("EROC_DerProviderName").distinct().count()
new_provider = number_of_providers - 142
pre_table_rows = 11 
copy_row = pre_table_rows + 142
end_row = copy_row + new_provider + 1

# copy and paste formatting onto unformatted rows
for row_number in range (copy_row, end_row):
    for column_number in range(2,end_column):
        cell_to_copy_from = ws_provider.cell(row=copy_row, column=column_number)
        cell_to_paste_to = ws_provider.cell(row= row_number, column=column_number)
        excel.copy_all_cell_styles(cell_to_copy_from, cell_to_paste_to)

# Use the values calculated above to get the cell range of the whole table
conditional_formatting_start_cell = "I11"
conditional_formatting_end_col = openpyxl.utils.cell.get_column_letter(end_column - 1)
conditional_formatting_end_row = end_row - 1
conditional_formatting_end_cell = conditional_formatting_end_col + str(conditional_formatting_end_row)
conditional_formatting_range = f"{conditional_formatting_start_cell}:{conditional_formatting_end_cell}"
print(conditional_formatting_range)

# Copy the existing conditional formatting rule, but make it cover the whole table using the range created above.
for rule_name, rule in ws_provider.conditional_formatting._cf_rules.items():
    ws_provider.conditional_formatting.add(conditional_formatting_range, rule[0])


# COMMAND ----------

#updating publishing date header
ws_provider.cell(row=3, column=3).value = date_header

# COMMAND ----------

from openpyxl.styles import NamedStyle

# Define the number format style
number_style = NamedStyle(name="number", number_format="0")

# Apply the number format to the specified range
for row in ws_provider.iter_rows(min_row=12, max_row=end_row, min_col=pre_date_columns + 1, max_col=end_column):
    for cell in row:
        cell.number_format = number_style.number_format

# COMMAND ----------

# MAGIC %md
# MAGIC #csv files
# MAGIC

# COMMAND ----------

#Month provider CSV
df_month_provider_csv = (df_processed_pifu
    .select (
        "EROC_DerProviderCode", 
        "EROC_DerProviderName", 
        "EROC_DerRegionName", 
        "EROC_DerRegionCode", 
        "EROC_DerICBCode", 
        "EROC_DerICBName", 
        "EROC_DerProviderAcuteStatus",
        "EROC_DerMonth", 
        "Moved_or_Discharged")
    .withColumnRenamed("EROC_DerProviderCode", "Provider Code")
    .withColumnRenamed("EROC_DerProviderName", "Provider Name")
    .withColumnRenamed("EROC_DerRegionName", "Region Name")
    .withColumnRenamed("EROC_DerRegionCode", "Region Code")
    .withColumnRenamed("EROC_DerICBCode", "ICB Code")
    .withColumnRenamed("EROC_DerICBName", "ICB Name")
    .withColumnRenamed("EROC_DerProviderAcuteStatus", "Acute Status")
    .withColumnRenamed("EROC_DerMonth", "Activity Month")
    .withColumnRenamed("Moved_or_Discharged", "Moved or Discharged")
)

df_month_provider_csv_pd = df_month_provider_csv.toPandas()

#specialty csv

df_specialty_csv = (df_tfc_pifu
    .select (
    "RTT_Specialty_code",
    "RTT_Specialty_Description",
    "EROC_DerMonth",
    "Moved_or_Discharged")
    .withColumnRenamed("RTT_Specialty_code", "Specialty Code")
    .withColumnRenamed("RTT_Specialty_Description", "Specialty Description")
    .withColumnRenamed("EROC_DerMonth", "Activity Month")
    .withColumnRenamed("Moved_or_Discharged", "Moved or Discharged")
)

df_specialty_csv_pd = df_specialty_csv.toPandas()

#save
df_month_provider_csv_pd.to_csv('outputs/PIFU_MI_Month_Provider.csv', index=False)
df_specialty_csv_pd.to_csv('outputs/PIFU_MI_Specialty.csv', index=False)

# COMMAND ----------

# MAGIC %md
# MAGIC #save report

# COMMAND ----------

# Save the workbook with the DataFrame inserted
wb.save('outputs/PIFU_MI.xlsx')
